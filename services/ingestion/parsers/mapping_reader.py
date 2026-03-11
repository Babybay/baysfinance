"""
Reads the 10_MAPPING_AKUN sheet from the Excel template.
Builds a mapping of sheet → column → account code configuration.
"""

from dataclasses import dataclass
from typing import Optional
from openpyxl.worksheet.worksheet import Worksheet


@dataclass
class MappingEntry:
    col_name: str           # e.g., "FOOD"
    account_code: str       # e.g., "4101"
    account_name: str       # e.g., "Pendapatan Makanan"
    side: str               # "DEBIT", "CREDIT", or "SKIP"
    description: str        # KETERANGAN


def read_mapping(workbook) -> dict[str, list[MappingEntry]]:
    """
    Read sheet '10_MAPPING_AKUN' starting at row 4.
    
    Expected columns:
    A: SHEET        (sheet name, e.g. "1_LAPORAN_PENJUALAN")
    B: KOLOM SUMBER (column name, e.g. "FOOD")
    C: KODE AKUN    (account code, e.g. "4101")
    D: NAMA AKUN    (account name, e.g. "Pendapatan Makanan")
    E: POSISI       (side: "DEBIT", "CREDIT", "SKIP")
    F: KETERANGAN   (notes)
    """
    sheet_name = "10_MAPPING_AKUN"
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' tidak ditemukan dalam file Excel.")

    ws: Worksheet = workbook[sheet_name]
    mapping: dict[str, list[MappingEntry]] = {}

    for row_num in range(4, ws.max_row + 1):
        sheet = _cell_str(ws, row_num, 1)  # col A
        col_name = _cell_str(ws, row_num, 2)  # col B
        account_code = _cell_str(ws, row_num, 3)  # col C
        account_name = _cell_str(ws, row_num, 4)  # col D
        side = _cell_str(ws, row_num, 5).upper()  # col E
        description = _cell_str(ws, row_num, 6)  # col F

        # Skip empty rows
        if not sheet or not col_name:
            continue

        # Normalize side
        if side not in ("DEBIT", "CREDIT", "SKIP"):
            if side in ("D", "DR"):
                side = "DEBIT"
            elif side in ("C", "CR"):
                side = "CREDIT"
            elif side in ("S", "-", "—", ""):
                side = "SKIP"

        entry = MappingEntry(
            col_name=col_name.strip(),
            account_code=account_code.strip() if account_code else "",
            account_name=account_name.strip() if account_name else "",
            side=side,
            description=description,
        )

        if sheet not in mapping:
            mapping[sheet] = []
        mapping[sheet].append(entry)

    return mapping


def get_column_index(ws: Worksheet, header_row: int, col_name: str) -> Optional[int]:
    """
    Find the column index (1-based) for a given column name in the header row.
    Uses fuzzy matching (case-insensitive, ignores extra spaces).
    """
    target = col_name.strip().upper()
    for col in range(1, ws.max_column + 1):
        cell_val = _cell_str(ws, header_row, col).upper()
        if cell_val == target:
            return col
        # Fuzzy: check if target is contained in header or vice versa
        if target in cell_val or cell_val in target:
            if len(target) > 3 and len(cell_val) > 3:
                return col
    return None


def build_column_map(
    ws: Worksheet, header_row: int, mapping_entries: list[MappingEntry]
) -> dict[str, int]:
    """
    Given the mapping entries for a sheet and its header row,
    build a map of col_name → column_index for all mapped columns.
    """
    col_map = {}
    for entry in mapping_entries:
        idx = get_column_index(ws, header_row, entry.col_name)
        if idx is not None:
            col_map[entry.col_name] = idx
    return col_map


def _cell_str(ws: Worksheet, row: int, col: int) -> str:
    """Get cell value as stripped string, or empty string if None."""
    val = ws.cell(row=row, column=col).value
    return str(val).strip() if val is not None else ""
