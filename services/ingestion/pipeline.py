"""
Pipeline orchestrator — reads an Excel template file and ingests all sheets
into the database as JournalEntries, FixedAssets, and FinancialReportSnapshots.
"""

import os
import traceback
from dataclasses import dataclass, field
from datetime import date
from typing import Optional
from openpyxl import load_workbook

from parsers.mapping_reader import read_mapping
from parsers.number_parser import parse_date
from parsers.sheet_parsers import (
    SheetResult,
    parse_laporan_penjualan,
    parse_piutang_lain_lain,
    parse_aset_tetap,
    parse_biaya_pra_operasi,
    parse_hutang_usaha,
    parse_hutang_owner,
    parse_laba_rugi,
    parse_arus_kas,
)
from db import (
    get_connection,
    create_import_batch,
    update_import_batch,
    insert_journal_entry,
    insert_fixed_asset,
    upsert_financial_snapshot,
    check_duplicate_entry,
    rollback_batch,
    get_import_batches,
)


REQUIRED_SHEETS = [
    "1_LAPORAN_PENJUALAN",
    "2_PIUTANG_LAIN_LAIN",
    "3_ASET_MANAJEMEN",
    "4_ASET_OWNER",
    "5_BIAYA_PRA_OPERASI",
    "6_HUTANG_USAHA",
    "7_HUTANG_OWNER",
    "8_LABA_RUGI",
    "9_ARUS_KAS",
]


@dataclass
class IngestionResult:
    batch_id: str
    status: str = "COMPLETED"
    company_name: str = ""
    period: str = ""
    total_entries: int = 0
    total_assets: int = 0
    total_skipped: int = 0
    total_errors: int = 0
    sheet_results: dict = field(default_factory=dict)
    warnings: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)


def ingest_template_file(
    file_path: str,
    client_id: str,
    imported_by: str = "System",
) -> IngestionResult:
    """
    Main orchestrator: reads an Excel template and ingests all sheets.
    
    Steps:
    1. Load workbook with data_only=True (resolve formulas)
    2. Validate required sheets
    3. Read company identity and period
    4. Read column→account mapping from sheet 10
    5. Process each sheet, creating JournalEntries / FixedAssets / Snapshots
    6. Track results in ImportBatch
    """
    conn = get_connection()
    batch_id = ""

    try:
        wb = load_workbook(file_path, data_only=True)

        # ── Step 2: Validate required sheets ──
        missing = [s for s in REQUIRED_SHEETS if s not in wb.sheetnames]
        if missing:
            raise ValueError(f"Sheet tidak ditemukan: {', '.join(missing)}")

        # ── Step 3: Read identity ──
        company_name, period = _read_identity(wb)

        # ── Step 4: Read mapping ──
        mapping = {}
        if "10_MAPPING_AKUN" in wb.sheetnames:
            mapping = read_mapping(wb)

        # ── Step 5: Create import batch ──
        batch_id = create_import_batch(
            conn, client_id, os.path.basename(file_path),
            period, company_name, imported_by,
        )
        conn.commit()

        result = IngestionResult(batch_id=batch_id, company_name=company_name, period=period)

        # ── Step 6: Process each sheet ──
        sheet_processors = [
            ("1_LAPORAN_PENJUALAN", lambda: parse_laporan_penjualan(
                wb["1_LAPORAN_PENJUALAN"], mapping.get("1_LAPORAN_PENJUALAN", [])
            )),
            ("2_PIUTANG_LAIN_LAIN", lambda: parse_piutang_lain_lain(wb["2_PIUTANG_LAIN_LAIN"])),
            ("3_ASET_MANAJEMEN", lambda: parse_aset_tetap(wb["3_ASET_MANAJEMEN"], "ASET_MANAJEMEN")),
            ("4_ASET_OWNER", lambda: parse_aset_tetap(wb["4_ASET_OWNER"], "ASET_OWNER")),
            ("5_BIAYA_PRA_OPERASI", lambda: parse_biaya_pra_operasi(wb["5_BIAYA_PRA_OPERASI"])),
            ("6_HUTANG_USAHA", lambda: parse_hutang_usaha(wb["6_HUTANG_USAHA"])),
            ("7_HUTANG_OWNER", lambda: parse_hutang_owner(wb["7_HUTANG_OWNER"])),
            ("8_LABA_RUGI", lambda: parse_laba_rugi(wb["8_LABA_RUGI"])),
            ("9_ARUS_KAS", lambda: parse_arus_kas(wb["9_ARUS_KAS"])),
        ]

        for sheet_name, processor in sheet_processors:
            try:
                sheet_result = processor()
                _persist_sheet_result(conn, sheet_result, client_id, batch_id, period)
                result.sheet_results[sheet_name] = {
                    "entries": len(sheet_result.entries),
                    "assets": len(sheet_result.assets),
                    "rows_processed": sheet_result.rows_processed,
                    "rows_skipped": sheet_result.rows_skipped,
                    "warnings": sheet_result.warnings,
                    "errors": sheet_result.errors,
                    "has_snapshot": sheet_result.snapshot is not None,
                }
                result.total_entries += len(sheet_result.entries)
                result.total_assets += len(sheet_result.assets)
                result.total_skipped += sheet_result.rows_skipped
                result.warnings.extend(sheet_result.warnings)

                conn.commit()

            except Exception as e:
                error_msg = f"Sheet {sheet_name}: {str(e)}"
                result.errors.append(error_msg)
                result.total_errors += 1
                conn.rollback()
                traceback.print_exc()

        # ── Step 7: Finalize batch ──
        final_status = "COMPLETED" if not result.errors else "COMPLETED"
        if result.total_entries == 0 and result.total_assets == 0:
            final_status = "FAILED"
            result.status = "FAILED"

        update_import_batch(
            conn, batch_id, final_status,
            entries_count=result.total_entries,
            skipped_count=result.total_skipped,
            errors_count=result.total_errors,
            warnings=result.warnings if result.warnings else None,
        )
        conn.commit()

        result.status = final_status
        return result

    except Exception as e:
        if batch_id:
            try:
                update_import_batch(conn, batch_id, "FAILED", errors_count=1, warnings=[str(e)])
                conn.commit()
            except Exception:
                conn.rollback()

        traceback.print_exc()
        return IngestionResult(
            batch_id=batch_id or "unknown",
            status="FAILED",
            errors=[str(e)],
        )
    finally:
        conn.close()


def _read_identity(wb) -> tuple[str, str]:
    """Read company name and period from any data sheet (B2, B3 or B4)."""
    company = ""
    period = ""

    for sheet_name in REQUIRED_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        val_b2 = ws.cell(row=2, column=2).value
        val_b3 = ws.cell(row=3, column=2).value
        val_b4 = ws.cell(row=4, column=2).value

        if val_b2 and not company:
            company = str(val_b2).strip()
        if val_b3 and not period:
            period = str(val_b3).strip()
        elif val_b4 and not period:
            period = str(val_b4).strip()

        if company and period:
            break

    return company, period


def _persist_sheet_result(
    conn, sheet_result: SheetResult, client_id: str, batch_id: str, period: str
):
    """Persist parsed sheet results to the database."""
    # Insert journal entries
    for entry_data in sheet_result.entries:
        # Duplicate check
        if check_duplicate_entry(conn, client_id, entry_data.source, entry_data.entry_date):
            sheet_result.warnings.append(
                f"Duplikat: {entry_data.source} tanggal {entry_data.entry_date} — dilewati"
            )
            sheet_result.rows_skipped += 1
            continue

        items = [
            {
                "account_code": item.account_code,
                "account_name": item.account_name,
                "debit": item.debit,
                "credit": item.credit,
            }
            for item in entry_data.items
            if item.debit > 0 or item.credit > 0
        ]

        if items:
            insert_journal_entry(
                conn, client_id, batch_id,
                entry_data.entry_date, entry_data.description,
                entry_data.source, items,
            )

    # Insert fixed assets
    for asset_data in sheet_result.assets:
        insert_fixed_asset(conn, client_id, batch_id, {
            "source": asset_data.source,
            "name": asset_data.name,
            "acquisition_date": asset_data.acquisition_date,
            "quantity": asset_data.quantity,
            "depreciation_rate": asset_data.depreciation_rate,
            "cost_prev": asset_data.cost_prev,
            "mutasi_in": asset_data.mutasi_in,
            "mutasi_out": asset_data.mutasi_out,
            "cost_current": asset_data.cost_current,
            "accum_deprec_prev": asset_data.accum_deprec_prev,
            "deprec_current_in": asset_data.deprec_current_in,
            "deprec_current_out": asset_data.deprec_current_out,
            "accum_deprec_current": asset_data.accum_deprec_current,
            "book_value": asset_data.book_value,
        })

    # Insert financial snapshots
    if sheet_result.snapshot and sheet_result.snapshot_type:
        upsert_financial_snapshot(
            conn, client_id, period,
            sheet_result.snapshot_type, sheet_result.snapshot,
        )
